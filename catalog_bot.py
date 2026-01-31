"""
Updated Telegram bot for selling rolled metal products with delivery.

This version extends the original ``catalog_bot.py`` implementation to
support three distinct user roles: administrators, suppliers and buyers
(clients).  Each role has its own menu and permitted actions.  The code
is intended as a starting point for the more complete specification
outlined in the project requirements.

Key additions compared to the original bot:

* Support for suppliers via the environment variable ``SUPPLIER_TG_IDS``.
  Telegram user IDs listed in this variable (comma‑separated) are
  considered suppliers.  Suppliers can view and update their own
  orders but cannot create new products or import price lists.
* Role detection helpers ``is_admin``, ``is_supplier`` and
  ``is_client`` to route incoming messages to the appropriate handlers.
* Separate reply keyboards for clients (buyers), suppliers and
  administrators.  Suppliers see a simplified menu with access to
  their orders and reports.
* Skeleton functions for supplier flows.  These handlers currently
  reuse the existing order listing logic but can be extended to
  implement editing order items, confirming shipments and other
  supplier actions as described in the functional specification.

This file is *not* a drop‑in replacement for the production bot.
It serves as a guide to help developers transition from a simple
admin/client model to a more sophisticated RBAC approach.  Additional
database migrations and business logic will be required to fully
implement the multi‑role state machine described in the specification.
"""

import os
import logging
import uuid
from decimal import Decimal
from collections import defaultdict
from datetime import datetime

import csv
try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:
    # openpyxl is optional; will be imported at runtime in handle_document if needed
    load_workbook = None  # type: ignore

from telegram import (
    Update,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

# Import database functions from shop_db.  Ensure that DATABASE_URL is set
# before running the bot.  shop_db.init_db() will create all necessary
# tables on startup.
from shop_db import (
    init_db,
    get_client_by_tg_id,
    insert_client,
    list_products,
    get_product_by_code,
    create_order,
    add_order_item,
    update_order_total,
    list_orders,
    get_order,
    set_order_status,
    record_payment,
    replace_products,
    list_orders_by_client,
    upsert_product,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
log = logging.getLogger("catalog_bot_updated")


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

TELEGRAM_BOT_TOKEN = os.getenv("8067476607:AAEhhNL6YISLFR9cj0ZUYquwkeI3FNFZAl8", "").strip()
ADMIN_TG_IDS_RAW = os.getenv("5084734764", "").strip()
SUPPLIER_TG_IDS_RAW = os.getenv("SUPPLIER_TG_IDS", "").strip()

if not TELEGRAM_BOT_TOKEN:
    raise RuntimeError("TELEGRAM_BOT_TOKEN is not set")

ADMIN_IDS: set[int] = set()
SUPPLIER_IDS: set[int] = set()

for part in (ADMIN_TG_IDS_RAW or "").split(","):
    part = part.strip()
    if part.isdigit():
        ADMIN_IDS.add(int(part))

for part in (SUPPLIER_TG_IDS_RAW or "").split(","):
    part = part.strip()
    if part.isdigit():
        SUPPLIER_IDS.add(int(part))


def is_admin(user_id: int) -> bool:
    """Return True if the given Telegram user ID belongs to an admin."""
    return user_id in ADMIN_IDS


def is_supplier(user_id: int) -> bool:
    """Return True if the given Telegram user ID belongs to a supplier."""
    return user_id in SUPPLIER_IDS


def is_client(user_id: int) -> bool:
    """
    Return True if the given Telegram user ID is neither admin nor supplier.
    Clients correspond to buyers in the business specification.
    """
    return not (is_admin(user_id) or is_supplier(user_id))


# ---------------------------------------------------------------------------
# Client and supplier session state
# ---------------------------------------------------------------------------

# Each client has a state dict storing the current step, cart and temp values.
# The structure is: { user_id: {"step": str | None, "cart": {code: qty},
#                             "pending_product": code | None,
#                             "address": str | None} }
CLIENT_STATE: dict[int, dict] = defaultdict(lambda: {"step": None, "cart": {}, "pending_product": None, "address": None})

# Suppliers may also require per‑user state for tracking edits or shipments.
SUPPLIER_STATE: dict[int, dict] = defaultdict(lambda: {"step": None, "pending_order": None})


# ---------------------------------------------------------------------------
# Keyboards
# ---------------------------------------------------------------------------

def client_menu_kb() -> ReplyKeyboardMarkup:
    """Main menu for clients (buyers)."""
    return ReplyKeyboardMarkup(
        [["🛍️ Каталог", "🛒 Корзина", "📦 Мои заказы"]],
        resize_keyboard=True,
    )


def supplier_menu_kb() -> ReplyKeyboardMarkup:
    """Main menu for suppliers."""
    return ReplyKeyboardMarkup(
        [["📦 Заявки", "📊 Отчёт"]],
        resize_keyboard=True,
    )


def admin_menu_kb() -> ReplyKeyboardMarkup:
    """Main menu for admins."""
    return ReplyKeyboardMarkup(
        [
            ["📦 Заказы", "📚 Товары", "➕ Товар"],
            ["📥 Импорт прайса", "📊 Отчёт"],
        ],
        resize_keyboard=True,
    )


def products_kb() -> InlineKeyboardMarkup:
    """Returns inline keyboard listing all products."""
    products = list_products(limit=100, offset=0)
    rows = []
    for p in products:
        code = p["code"]
        name = p["name"]
        price = p["price"]
        unit = p.get("unit", "")
        label = f"{name} — {price:g}{('/' + unit) if unit else ''}"
        rows.append([InlineKeyboardButton(label, callback_data=f"prod:{code}")])
    if not rows:
        rows.append([InlineKeyboardButton("Каталог пуст", callback_data="noop")])
    return InlineKeyboardMarkup(rows)


def cart_kb(has_items: bool) -> InlineKeyboardMarkup:
    """Returns inline keyboard for cart actions."""
    rows = []
    if has_items:
        rows.append([InlineKeyboardButton("📦 Оформить заказ", callback_data="cart:place")])
        rows.append([InlineKeyboardButton("🧹 Очистить", callback_data="cart:clear")])
    rows.append([InlineKeyboardButton("🔙 В каталог", callback_data="cart:back")])
    return InlineKeyboardMarkup(rows)


def admin_orders_kb(orders) -> InlineKeyboardMarkup:
    """Returns inline keyboard for admin orders list."""
    rows = []
    for o in orders:
        order_id = o["id"]
        status = o["status"]
        total = o["total_amount"]
        label = f"{order_id[:8]}… | {total:g} | {status}"
        rows.append([InlineKeyboardButton(label, callback_data=f"order:{order_id}")])
    if not rows:
        rows.append([InlineKeyboardButton("Нет заказов", callback_data="noop")])
    return InlineKeyboardMarkup(rows)


def supplier_orders_kb(orders) -> InlineKeyboardMarkup:
    """Returns inline keyboard for supplier orders list."""
    # For now, reuse the admin layout.  Suppliers may later need to see
    # only their own orders or have different actions available.
    return admin_orders_kb(orders)

# ---------------------------------------------------------------------------
# Order status keyboard
# ---------------------------------------------------------------------------

def order_status_kb(order_id: str) -> InlineKeyboardMarkup:
    """Return an inline keyboard for updating order status.

    The order status workflow described in the business requirements involves
    multiple states beyond the original ``new`` → ``shipped`` → ``delivered`` →
    ``paid`` model.  To support a richer state machine while maintaining
    backwards compatibility, this helper constructs a keyboard with all
    supported statuses.  When a button is pressed, a callback payload of
    ``setstat:<order_id>:<status_code>`` will be sent to the bot, which the
    ``on_callback`` handler uses to update the order via ``set_order_status``.

    ``order_id`` should be the unique identifier of the order.  The button
    labels are localized Russian names for readability.  You can reorder or
    prune the list as needed, but ensure that the callback data values
    (status codes) match what your backend expects.
    """
    # Define the list of statuses as tuples (internal_code, label)
    statuses: list[tuple[str, str]] = [
        ("draft", "Черновик"),
        ("submitted", "Отправлено"),
        ("under_review", "На рассмотрении"),
        ("needs_approval", "Требует согласования"),
        ("agreed", "Согласовано"),
        ("confirmed", "Подтверждено"),
        ("shipped", "Отгружено"),
        ("received", "Получено"),
        ("paid", "Оплачено"),
        ("closed", "Закрыто"),
        ("cancelled", "Отменено"),
    ]
    rows: list[list[InlineKeyboardButton]] = []
    for status_code, label in statuses:
        callback_data = f"setstat:{order_id}:{status_code}"
        rows.append([InlineKeyboardButton(label, callback_data=callback_data)])
    return InlineKeyboardMarkup(rows)


# ---------------------------------------------------------------------------
# Command and message handlers
# ---------------------------------------------------------------------------

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /start command for all roles."""
    user = update.effective_user
    if not user:
        return
    uid = user.id
    # Determine role and present appropriate menu
    if is_admin(uid):
        await update.message.reply_text(
            "Меню администратора", reply_markup=admin_menu_kb()
        )
        return
    if is_supplier(uid):
        await update.message.reply_text(
            "Меню поставщика", reply_markup=supplier_menu_kb()
        )
        return
    # client flow
    client = get_client_by_tg_id(uid)
    if client:
        await update.message.reply_text(
            "Главное меню", reply_markup=client_menu_kb()
        )
    else:
        # ask for phone number
        contact_btn = KeyboardButton("📲 Отправить номер телефона", request_contact=True)
        markup = ReplyKeyboardMarkup([[contact_btn]], resize_keyboard=True)
        await update.message.reply_text(
            "Здравствуйте! Чтобы оформить заказ, пожалуйста, поделитесь своим номером телефона.",
            reply_markup=markup,
        )


async def handle_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle contact message from client."""
    user = update.effective_user
    if not user or not is_client(user.id):
        return
    contact = update.message.contact
    if not contact:
        return
    phone = contact.phone_number
    name = f"{contact.first_name or ''} {contact.last_name or ''}".strip()
    client_id = insert_client(tg_id=user.id, phone=phone, name=name)
    log.info("Registered client %s with id %s", phone, client_id)
    # reset client state
    CLIENT_STATE[user.id] = {"step": None, "cart": {}, "pending_product": None, "address": None}
    await update.message.reply_text(
        "Спасибо! Теперь вы можете оформить заказ.", reply_markup=client_menu_kb()
    )


async def on_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle text messages for all roles."""
    user = update.effective_user
    if not user:
        return
    uid = user.id
    text = (update.message.text or "").strip()
    # Admin commands
    if is_admin(uid):
        await handle_admin_text(update, context, text)
        return
    # Supplier commands
    if is_supplier(uid):
        await handle_supplier_text(update, context, text)
        return
    # Client commands
    state = CLIENT_STATE[uid]
    if state.get("step") == "enter_qty":
        # expecting quantity for product
        await client_receive_quantity(update, context)
        return
    if state.get("step") == "enter_address":
        # expecting address for order
        await client_receive_address(update, context)
        return
    # menu actions for client
    if text == "🛍️ Каталог":
        await client_show_catalog(update, context)
        return
    if text == "🛒 Корзина":
        await client_show_cart(update, context)
        return
    if text == "📦 Мои заказы":
        await client_show_orders(update, context)
        return
    await update.message.reply_text(
        "Выберите действие через меню.", reply_markup=client_menu_kb()
    )


async def handle_admin_text(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
    """Handle admin commands via text menu."""
    # If admin is in the middle of adding a product
    if context.user_data.get("add_product_step"):
        await admin_handle_add_product(update, context, text)
        return
    # Admin menu actions
    if text == "📦 Заказы":
        orders = list_orders(limit=50, offset=0)
        kb = admin_orders_kb(orders)
        await update.message.reply_text("Список заказов:", reply_markup=kb)
        return
    if text == "📚 Товары":
        kb = products_kb()
        await update.message.reply_text("Каталог товаров:", reply_markup=kb)
        return
    if text == "➕ Товар":
        context.user_data["add_product_step"] = "code"
        context.user_data["add_product_data"] = {}
        await update.message.reply_text("Введите код товара:")
        return
    if text == "📊 Отчёт":
        await admin_show_report(update, context)
        return
    if text == "📥 Импорт прайса":
        # prompt admin to send an Excel or CSV file
        await update.message.reply_text(
            "Отправьте Excel‑файл (.xlsx, .xlsm) или CSV с колонками code, name, price, unit (необязательно), description (необязательно).",
        )
        # set admin state to expect file
        context.user_data["awaiting_price_file"] = True
        return
    await update.message.reply_text(
        "Используйте меню для выбора действия.", reply_markup=admin_menu_kb()
    )


async def handle_supplier_text(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
    """Handle supplier commands via text menu.

    This is a simplified skeleton that mirrors the admin order view.  In a
    full implementation, suppliers should only see orders assigned to them
    and should have capabilities to confirm/edit those orders and mark
    shipments or deliveries.  Additional state management may be required
    to support multi‑step interactions such as editing order items.
    """
    if text == "📦 Заявки":
        # Suppliers currently see all orders.  Filter by supplier in a real app.
        orders = list_orders(limit=50, offset=0)
        kb = supplier_orders_kb(orders)
        await update.message.reply_text("Список заявок:", reply_markup=kb)
        return
    if text == "📊 Отчёт":
        # Reuse admin report for now.  In the future, aggregate only supplier orders.
        await admin_show_report(update, context)
        return
    await update.message.reply_text(
        "Используйте меню для выбора действия.", reply_markup=supplier_menu_kb()
    )


# ---------------------------------------------------------------------------
# Document handler (price import) – remains admin only
# ---------------------------------------------------------------------------

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle document messages (used by admin for price import)."""
    user = update.effective_user
    if not user or not is_admin(user.id):
        return
    if not context.user_data.get("awaiting_price_file"):
        return
    doc = update.message.document
    if not doc:
        return
    file_name = doc.file_name or ""
    # We accept Excel (.xlsx, .xlsm) or CSV (.csv) files
    ext = file_name.lower().split(".")[-1]
    allowed_exts = {"xlsx", "xlsm", "csv"}
    if ext not in allowed_exts:
        await update.message.reply_text("Пожалуйста, отправьте файл Excel (.xlsx, .xlsm) или CSV (.csv).")
        return
    # download file to a temporary location
    tmp_path = f"/tmp/{uuid.uuid4()}_{file_name}"
    file_obj = await doc.get_file()
    await file_obj.download_to_drive(tmp_path)
    items: list[dict] = []
    try:
        if ext == "csv":
            # Parse CSV file using built‑in csv module
            with open(tmp_path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    code = str(row.get("code", "")).strip()
                    name = str(row.get("name", "")).strip()
                    if not code or not name:
                        continue
                    price = row.get("price", 0)
                    try:
                        price_val = float(price) if price not in (None, "", "nan") else 0.0
                    except Exception:
                        price_val = 0.0
                    unit = str(row.get("unit", "") or "").strip()
                    desc = str(row.get("description", "") or "").strip()
                    items.append({"code": code, "name": name, "price": price_val, "unit": unit, "description": desc})
        else:
            # Parse Excel file using openpyxl
            if load_workbook is None:
                raise ImportError("openpyxl is required to parse Excel files but is not installed")
            wb = load_workbook(tmp_path, data_only=True)
            sheet = wb.active
            # read header row (first row)
            rows_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            headers_row = next(rows_iter, None)
            if not headers_row:
                raise ValueError("Файл пуст или не содержит заголовков")
            headers = [str(h).strip().lower() if h is not None else "" for h in headers_row]
            header_index = {name: idx for idx, name in enumerate(headers) if name}
            # iterate over the remaining rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                code = str(row[header_index.get("code", -1)] or "").strip() if "code" in header_index else ""
                name = str(row[header_index.get("name", -1)] or "").strip() if "name" in header_index else ""
                if not code or not name:
                    continue
                price_val = 0.0
                if "price" in header_index:
                    cell_val = row[header_index["price"]]
                    try:
                        price_val = float(cell_val) if cell_val not in (None, "", "nan") else 0.0
                    except Exception:
                        price_val = 0.0
                unit = str(row[header_index.get("unit", -1)] or "").strip() if "unit" in header_index else ""
                desc = str(row[header_index.get("description", -1)] or "").strip() if "description" in header_index else ""
                items.append({"code": code, "name": name, "price": price_val, "unit": unit, "description": desc})
    except Exception as e:
        log.exception("Failed to read price file: %s", e)
        await update.message.reply_text(f"Не удалось прочитать файл: {e}")
        context.user_data["awaiting_price_file"] = False
        return
    if not items:
        await update.message.reply_text("В файле нет валидных строк.")
    else:
        replace_products(items)
        await update.message.reply_text(f"Импортировано {len(items)} записей.", reply_markup=admin_menu_kb())
    context.user_data["awaiting_price_file"] = False


# ---------------------------------------------------------------------------
# Client flow functions (unchanged)
# ---------------------------------------------------------------------------

async def client_show_catalog(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show catalog of products to the client."""
    kb = products_kb()
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text("Выберите товар:", reply_markup=kb)
    else:
        await update.message.reply_text("Выберите товар:", reply_markup=kb)


async def client_show_cart(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Display cart summary to client."""
    user_id = update.effective_user.id
    cart = CLIENT_STATE[user_id]["cart"]
    if not cart:
        msg = "Корзина пуста."
    else:
        lines = ["Ваша корзина:"]
        total = Decimal("0")
        for code, qty in cart.items():
            prod = get_product_by_code(code)
            if not prod:
                continue
            name = prod["name"]
            price = Decimal(str(prod["price"]))
            amount = price * Decimal(str(qty))
            total += amount
            unit = prod.get("unit", "")
            lines.append(f"• {name} — {qty:g}{('/' + unit) if unit else ''} × {price:g} = {amount:g}")
        lines.append(f"\nИтого: {total:g}")
        msg = "\n".join(lines)
    kb = cart_kb(bool(cart))
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(msg, reply_markup=kb)
    else:
        await update.message.reply_text(msg, reply_markup=kb)


async def client_show_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show orders history to the client."""
    user_id = update.effective_user.id
    client = get_client_by_tg_id(user_id)
    if not client:
        await update.message.reply_text("Сначала зарегистрируйтесь, поделившись номером телефона.")
        return
    orders = list_orders_by_client(client["id"])
    if not orders:
        await update.message.reply_text("У вас пока нет заказов.")
        return
    lines = ["Ваши заказы:"]
    for o in orders:
        order_id = o["id"]
        created = o["created_at"].strftime("%d.%m.%Y") if o["created_at"] else "?"
        status = o["status"]
        total = o["total_amount"]
        lines.append(f"• {order_id[:8]}… | {created} | {total:g} | {status}")
    await update.message.reply_text("\n".join(lines))


async def client_receive_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle quantity input from client."""
    user_id = update.effective_user.id
    state = CLIENT_STATE[user_id]
    text_val = (update.message.text or "").replace(",", ".").strip()
    try:
        qty = float(text_val)
        if qty <= 0:
            raise ValueError
    except Exception:
        await update.message.reply_text("Введите корректное количество, например 2.5")
        return
    code = state.get("pending_product")
    if not code:
        await update.message.reply_text("Произошла ошибка. Попробуйте снова.")
        state["step"] = None
        return
    # add to cart
    cart = state["cart"]
    cart[code] = cart.get(code, 0) + qty
    state["step"] = None
    state["pending_product"] = None
    # prompt next action
    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("➕ Продолжить", callback_data="shop:more"),
            InlineKeyboardButton("🛒 Корзина", callback_data="cart:show"),
        ]
    ])
    await update.message.reply_text("Товар добавлен в корзину.", reply_markup=kb)


async def client_receive_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle address input from client and place order."""
    user_id = update.effective_user.id
    state = CLIENT_STATE[user_id]
    address = update.message.text.strip()
    if not address:
        await update.message.reply_text("Введите корректный адрес доставки.")
        return
    client = get_client_by_tg_id(user_id)
    if not client:
        await update.message.reply_text("Ошибка: клиент не найден. Попробуйте /start.")
        state["step"] = None
        return
    cart = state["cart"]
    if not cart:
        await update.message.reply_text("Ваша корзина пуста.")
        state["step"] = None
        return
    # create order
    order_id = create_order(client_id=client["id"])
    total_amount = Decimal("0")
    for code, qty in cart.items():
        prod = get_product_by_code(code)
        if not prod:
            continue
        price = Decimal(str(prod["price"]))
        amount = price * Decimal(str(qty))
        total_amount += amount
        add_order_item(order_id=order_id, product_id=prod["id"], quantity=qty, price=float(price))
    # update total
    update_order_total(order_id)
    # set initial status for the new order to submitted.  In the legacy
    # implementation ``create_order`` may set status to ``new``; here we
    # explicitly update it to ``submitted`` to align with the extended state
    # machine defined in the functional specification.
    try:
        set_order_status(order_id, "submitted")
    except Exception:
        # Fallback silently if backend does not support this status yet
        pass
    # clear cart
    state["cart"] = {}
    state["step"] = None
    state["address"] = address
    # notify client
    await update.message.reply_text(
        f"Заказ создан! Номер заказа: {order_id}.\nНаш менеджер свяжется с вами для подтверждения. Спасибо за заказ!",
        reply_markup=client_menu_kb(),
    )
    # notify suppliers (for now notify all suppliers).  In a complete implementation,
    # orders would be routed to a specific supplier based on the selected supplier or product.
    order_info_lines = [f"📦 Новый заказ {order_id}"]
    order_info_lines.append(f"Клиент: {client.get('name','')} / {client.get('phone','')}")
    order_info_lines.append(f"Адрес: {address}")
    order_info_lines.append("Состав заказа:")
    for code, qty in cart.items():
        prod = get_product_by_code(code)
        if not prod:
            continue
        name = prod["name"]
        unit = prod.get("unit", "")
        order_info_lines.append(f"• {name} — {qty:g}{('/' + unit) if unit else ''}")
    order_info_lines.append(f"Итого: {total_amount:g}")
    info = "\n".join(order_info_lines)
    # send to all suppliers
    for sup_id in SUPPLIER_IDS:
        try:
            await context.bot.send_message(chat_id=sup_id, text=info)
        except Exception as e:
            log.exception("Failed to notify supplier %s: %s", sup_id, e)


# ---------------------------------------------------------------------------
# Callback query handler (client and admin flows).  Supplier callbacks can
# reuse these handlers for now.  Extend as needed for supplier actions.
# ---------------------------------------------------------------------------

async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle callback queries for all roles."""
    query = update.callback_query
    if not query:
        return
    user = update.effective_user
    data = query.data or ""
    uid = user.id if user else 0
    # Admin callbacks
    if is_admin(uid):
        if data.startswith("order:"):
            _, oid = data.split(":", 1)
            await admin_show_order(update, context, oid)
            return
        if data.startswith("setstat:"):
            _, oid, status = data.split(":", 2)
            set_order_status(oid, status)
            await query.answer("Статус обновлён")
            await admin_show_order(update, context, oid)
            return
        if data == "noop":
            await query.answer()
            return
    # Supplier callbacks – suppliers can view orders and update certain statuses.
    if is_supplier(uid):
        if data.startswith("order:"):
            _, oid = data.split(":", 1)
            await admin_show_order(update, context, oid)
            return
        if data.startswith("setstat:"):
            # allow suppliers to update only specific statuses
            try:
                _, oid, status = data.split(":", 2)
            except ValueError:
                await query.answer("Неверный формат команды", show_alert=True)
                return
            # Define which statuses suppliers are permitted to set
            allowed_statuses = {"confirmed", "shipped", "received"}
            if status not in allowed_statuses:
                await query.answer("Недопустимый статус", show_alert=True)
                return
            # Attempt to update the status
            try:
                set_order_status(oid, status)
            except Exception as exc:
                log.exception("Supplier failed to set status: %s", exc)
                await query.answer("Ошибка при обновлении статуса", show_alert=True)
                return
            await query.answer("Статус обновлён")
            await admin_show_order(update, context, oid)
            return
        if data == "noop":
            await query.answer()
            return
    # Client callbacks
    if data.startswith("prod:"):
        _, code = data.split(":", 1)
        await client_select_product(update, context, code)
        return
    if data == "shop:more":
        await client_show_catalog(update, context)
        return
    if data.startswith("cart:"):
        _, action = data.split(":", 1)
        if action == "show":
            await client_show_cart(update, context)
            return
        if action == "back":
            await client_show_catalog(update, context)
            return
        if action == "clear":
            CLIENT_STATE[uid]["cart"] = {}
            await query.answer("Корзина очищена")
            await client_show_cart(update, context)
            return
        if action == "place":
            # ask for address
            CLIENT_STATE[uid]["step"] = "enter_address"
            await query.answer()
            await query.edit_message_text(
                "Введите адрес доставки (улица, дом, комментарий):"
            )
            return
    if data == "noop":
        await query.answer()
        return


# ---------------------------------------------------------------------------
# Admin helper functions
# ---------------------------------------------------------------------------

async def admin_handle_add_product(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
    """Handle step‑by‑step product creation for admin.

    This function uses ``context.user_data['add_product_step']`` to track which field
    is being entered and ``context.user_data['add_product_data']`` to accumulate
    the partial product data. When all required fields have been collected, the
    product is inserted or updated via ``upsert_product`` and the admin is
    returned to the main menu.

    Steps:
      - ``code``: product code (unique identifier)
      - ``name``: product name
      - ``price``: numeric price per unit
      - ``unit``: unit of measurement (optional)
      - ``desc``: description (optional)
    """
    step = context.user_data.get("add_product_step")
    data = context.user_data.get("add_product_data", {})
    if step is None:
        # Something went wrong; reset state
        context.user_data.pop("add_product_step", None)
        context.user_data.pop("add_product_data", None)
        await update.message.reply_text(
            "Неверное состояние. Попробуйте снова выбрать \"➕ Товар\".",
            reply_markup=admin_menu_kb(),
        )
        return
    text_value = text.strip()
    if step == "code":
        if not text_value:
            await update.message.reply_text("Код не может быть пустым. Введите код товара:")
            return
        data["code"] = text_value
        context.user_data["add_product_step"] = "name"
        context.user_data["add_product_data"] = data
        await update.message.reply_text("Введите название товара:")
        return
    if step == "name":
        if not text_value:
            await update.message.reply_text("Название не может быть пустым. Введите название товара:")
            return
        data["name"] = text_value
        context.user_data["add_product_step"] = "price"
        context.user_data["add_product_data"] = data
        await update.message.reply_text("Введите цену за единицу (например, 25.5):")
        return
    if step == "price":
        text_norm = text_value.replace(",", ".")
        try:
            price_val = float(text_norm)
            if price_val < 0:
                raise ValueError
        except Exception:
            await update.message.reply_text("Введите корректную цену, например 10.5")
            return
        data["price"] = price_val
        context.user_data["add_product_step"] = "unit"
        context.user_data["add_product_data"] = data
        await update.message.reply_text(
            "Введите единицу измерения (например, кг, м; оставьте пустым, если не нужно):"
        )
        return
    if step == "unit":
        unit_val = text_value
        data["unit"] = unit_val
        context.user_data["add_product_step"] = "desc"
        context.user_data["add_product_data"] = data
        await update.message.reply_text("Введите описание (или '-' для пропуска):")
        return
    if step == "desc":
        desc_val = "" if text_value == "-" else text_value
        data["description"] = desc_val
        try:
            upsert_product(
                code=data.get("code"),
                name=data.get("name"),
                price=data.get("price"),
                unit=data.get("unit", ""),
                description=data.get("description", ""),
            )
        except Exception as exc:
            log.exception("Ошибка при добавлении товара: %s", exc)
            await update.message.reply_text(f"Не удалось добавить товар: {exc}")
        else:
            await update.message.reply_text(
                f"Товар '{data.get('name')}' добавлен/обновлён.", reply_markup=admin_menu_kb()
            )
        # Reset state
        context.user_data.pop("add_product_step", None)
        context.user_data.pop("add_product_data", None)
        return


async def admin_show_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Generate and send a summary report of orders and balances to the admin.

    The report includes total number of orders, number of paid and unpaid orders,
    the total amount of all orders, the total unpaid amount (debt), and a
    listing of unpaid orders with client names and amounts.
    """
    orders = list_orders(limit=1000, offset=0)
    if not orders:
        await update.message.reply_text("Нет заказов.", reply_markup=admin_menu_kb())
        return
    total_count = 0
    paid_count = 0
    unpaid_count = 0
    total_sum = Decimal("0")
    debt_sum = Decimal("0")
    unpaid_lines: list[str] = []
    for o in orders:
        total_count += 1
        amount = Decimal(str(o.get("total_amount") or 0))
        total_sum += amount
        status = o.get("status", "")
        if status == "paid":
            paid_count += 1
        else:
            unpaid_count += 1
            debt_sum += amount
            client_name = o.get("name") or o.get("phone") or "?"
            order_id = o["id"]
            unpaid_lines.append(
                f"• {order_id[:8]}… | {client_name} | {amount:g} | {status}"
            )
    lines: list[str] = []
    lines.append("📊 Отчёт")
    lines.append(f"Всего заказов: {total_count}")
    lines.append(f"Оплачено: {paid_count}")
    lines.append(f"Не оплачено: {unpaid_count}")
    lines.append(f"Общая сумма: {total_sum:g}")
    lines.append(f"Сумма задолженности: {debt_sum:g}")
    if unpaid_lines:
        lines.append("\nНеоплаченные заказы:")
        lines.extend(unpaid_lines)
    else:
        lines.append("\nВсе заказы оплачены")
    msg = "\n".join(lines)
    await update.message.reply_text(msg, reply_markup=admin_menu_kb())
    return


async def client_select_product(update: Update, context: ContextTypes.DEFAULT_TYPE, code: str):
    """Handle product selection by client."""
    user_id = update.effective_user.id
    prod = get_product_by_code(code)
    if not prod:
        await update.callback_query.answer("Товар не найден", show_alert=True)
        return
    name = prod["name"]
    price = prod["price"]
    unit = prod.get("unit", "")
    msg = f"{name}\nЦена: {price:g}{('/' + unit) if unit else ''}\nВведите количество:"
    # update state
    CLIENT_STATE[user_id]["step"] = "enter_qty"
    CLIENT_STATE[user_id]["pending_product"] = code
    await update.callback_query.answer()
    await update.callback_query.edit_message_text(msg)


async def admin_show_order(update: Update, context: ContextTypes.DEFAULT_TYPE, order_id: str):
    """Show order details to admin or supplier."""
    order = get_order(order_id)
    if not order:
        await update.callback_query.answer("Заказ не найден", show_alert=True)
        return
    lines: list[str] = []
    lines.append(f"🧾 Заказ {order['id']}")
    lines.append(f"Дата: {order['created_at'].strftime('%d.%m.%Y %H:%M') if order['created_at'] else ''}")
    client_info = f"{order.get('client_name','')} / {order.get('client_phone','')}"
    lines.append(f"Клиент: {client_info}")
    lines.append(f"Статус: {order['status']}")
    lines.append("\nПозиции:")
    for it in order['items']:
        qty = it['quantity']
        price = it['price']
        amount = it['amount']
        name = it.get('product_name', it['product_id'])
        unit = it.get('product_unit', '')
        lines.append(f"• {name} — {qty:g}{('/' + unit) if unit else ''} × {price:g} = {amount:g}")
    lines.append(f"\nИтого: {order['total_amount']:g}")
    msg = "\n".join(lines)
    # Choose appropriate keyboard: admins and suppliers can update status.
    # Present the status keyboard for both admin and supplier roles.  In client
    # context no status keyboard is shown.
    kb = None
    uid = update.effective_user.id if update.effective_user else 0
    if is_admin(uid) or is_supplier(uid):
        kb = order_status_kb(order_id)
    await update.callback_query.answer()
    await update.callback_query.edit_message_text(msg, reply_markup=kb)


# ---------------------------------------------------------------------------
# Bot startup
# ---------------------------------------------------------------------------

def main() -> None:
    """Entry point to run the bot."""
    log.info("Starting updated catalog bot…")
    # initialize database (creates tables and performs migrations)
    init_db()
    app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    # contact handler for clients
    app.add_handler(MessageHandler(filters.CONTACT, handle_contact))
    # document handler for admins (price import)
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    # text messages
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))
    # callback queries
    app.add_handler(CallbackQueryHandler(on_callback))
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
